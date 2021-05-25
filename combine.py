import os, re
from openpyxl import load_workbook
import pandas as pd
import numpy as np
#find all Excel files
def list_files(dir):
    r = []
    for root, dirs, files in os.walk(dir):
        for name in files:
            r.append(os.path.join(root, name))
    return r
# using Regex to filter result.
reg = re.compile(r'.*\\.*\\\d\_Application Form*\.(xlsx|xls|csv)$')

def filter_list(list):
    return [ s for s in list if reg.match(s) ]    

list = list_files("c:/Users/P/Desktop/Use Case 3/")

columns = ['Application ID', 'Name', 'Position', 'Company',	'E-mail', 'Phone Number', 'Payment Method','Price', 'Status']
emtpy_list = []

for file in filter_list(list):
     if file.endswith('.xlsx'):
         file_list = []
         workbook = load_workbook(filename=file, read_only=True)
         worksheet = workbook[workbook.sheetnames[0]]
         file_list.append(worksheet['D3'].value)
         file_list.append(worksheet['B3'].value)
         file_list.append(worksheet['C9'].value)
         file_list.append(worksheet['C6'].value)
         file_list.append(worksheet['C11'].value)
         file_list.append(worksheet['C10'].value)
         file_list.append(worksheet['B15'].value)
         file_list.append(worksheet['C15'].value)
         file_list.append(worksheet['D15'].value)
         emtpy_list.append(file_list)

# Transpose Data
data = np.array(emtpy_list).T.tolist()
# Create Dataframe
df = pd.DataFrame(data, columns).transpose()
print(df.head)
# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter('prabhdeep.singh@shinybluebox.com_Bot_3_Output.xlsx', engine='xlsxwriter')
# Convert the dataframe to an XlsxWriter Excel object.
df.to_excel(writer, sheet_name='Attendees', index=False)
# Close the Pandas Excel writer and output the Excel file.
writer.save()
